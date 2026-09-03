"""§CMT-1 货 5：写手说明、Excel 类型列、采集卡 comments 开关。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pytest

from app.adapters.capability import Capability
from app.adapters.source_mcp import codex_mcp_args, stdio_server_config
from app.api.delivery import EVIDENCE_FIELDS, evidence_view
from app.export.excel import KIND_LABELS, SOURCE_HEADER
from app.plan.generate import render_source_handbook


def test_采集卡默认开评论二跳_闭集只认_on_off() -> None:
    assert Capability().comments == "on"
    assert Capability(comments="off").comments == "off"
    with pytest.raises(ValueError, match="on/off"):
        Capability(comments="maybe")


def test_开关随_mcp_配置下发_默认不改配置形状() -> None:
    default = stdio_server_config(("xhs",))["args"]
    assert "--no-comments" not in default
    off = stdio_server_config(("xhs",), comments="off")["args"]
    assert off == [*default, "--no-comments"]
    assert any("--no-comments" in item for item in codex_mcp_args(("xhs",), comments="off"))


def test_命令行参数能把开关关掉() -> None:
    from app.adapters.source_mcp import _parser

    assert _parser().parse_args(["--source", "xhs"]).no_comments is False
    assert _parser().parse_args(["--source", "xhs", "--no-comments"]).no_comments is True


def test_采集卡生成时带上开关() -> None:
    from app.plan.generate import _capability

    card = _capability("web-collector", "goal-1", [], source_id="xhs")
    assert card["comments"] == "on"
    assert Capability(**{k: v for k, v in card.items() if k != "fs"}, fs=card["fs"])


def test_信息源手册讲了评论二跳且逐源表顺延到第六节() -> None:
    page = render_source_handbook("fast")
    assert "## 4. 评论是自动带上的" in page
    assert "capability.comments" in page
    assert "stance=contradicts" in page
    assert "## 6. 当前可用的源" in page


def _row(index: int, **overrides: Any) -> dict[str, Any]:
    row = {
        "id": f"ev-{index}", "citation_no": index,
        "permalink": f"https://www.xiaohongshu.com/explore/n{index}",
        "title": f"笔记 {index}", "content_excerpt": "正文", "platform": "xhs",
        "source_type": "post", "fetch_method": "third_party_api",
        "author_name": "作者", "published_at": None,
        "fetched_at": "2026-09-03T00:00:00+00:00", "goal_id": "goal-1",
        "score_authority": 1, "score_freshness": 2, "score_crossref": 1,
        "score_completeness": 1, "score_independence": 2, "score_total": 7,
        "grade": "B", "rated_by": "baseline:xhs@v1", "raw_metrics": {},
        "rating_notes": (
            "权威1:平台社区基线 · 时效2:请求时间窗限定 · 交叉1:弱源或已说明分歧"
            " · 完整1:正文摘要可回溯 · 无关2:无可见利益关系"
        ),
        "kind": "post", "parent_permalink": None,
    }
    row.update(overrides)
    return row


def _comment_row(index: int) -> dict[str, Any]:
    parent = "https://www.xiaohongshu.com/explore/n1"
    return _row(
        index, kind="comment", source_type="comment", parent_permalink=parent,
        permalink=f"{parent}?owli_comment=c{index}", title="评论 · 笔记 1",
        author_name=f"读者{index}",
        rating_notes=(
            "权威0:评论·作者不可核验 · 时效2:请求时间窗限定 · 交叉1:弱源或已说明分歧"
            " · 完整1:正文摘要可回溯 · 无关2:无可见利益关系"
        ),
        score_authority=0, score_total=6,
    )


def test_证据接口把_kind_与父帖链接透出并分类计数() -> None:
    assert {"kind", "parent_permalink"} <= set(EVIDENCE_FIELDS)
    view = evidence_view([_row(1), _comment_row(2)])
    assert view["counts"]["by_kind"] == {"post": 1, "comment": 1}
    assert view["items"][1]["kind"] == "comment"
    assert view["items"][1]["parent_permalink"] == (
        "https://www.xiaohongshu.com/explore/n1"
    )


def test_excel_信息源表有类型列且校验器认它(tmp_path: Path) -> None:
    from app.export.excel import build_workbook
    from app.export.excel_check import check_workbook

    assert SOURCE_HEADER[2] == "类型"
    assert KIND_LABELS == {"post": "帖", "comment": "评论"}
    report = {"id": "r-1", "title": "报告", "research_question": "问题"}
    view = {
        "conclusions": ["结论一 [S01][S02]"],
        "sources": [
            {"citation_no": 1, "permalink": _row(1)["permalink"], "title": "笔记 1"},
            {"citation_no": 2, "permalink": _comment_row(2)["permalink"],
             "title": "评论 · 笔记 1"},
        ],
    }
    workbook = build_workbook(report, view, [_row(1), _comment_row(2)], [])
    path = tmp_path / "out.xlsx"
    workbook.save(path)
    assert check_workbook(path) == []

    from openpyxl import load_workbook

    sheet = load_workbook(path)["04_信息源"]
    assert [sheet.cell(row=r, column=3).value for r in (2, 3)] == ["帖", "评论"]
