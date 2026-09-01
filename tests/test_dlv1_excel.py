"""§DLV-1 货 3：Excel 6 sheet 确定性导出 + 校验器（尺子自己也要验）。"""

from __future__ import annotations

import zipfile
from pathlib import Path

from openpyxl import load_workbook

from tests.test_dlv1_delivery import _seed_evidence, _seed_history, _write_json_report


def _export(tmp_path: Path) -> tuple[Path, str]:
    from app.export.excel import export_excel
    from app.store.dao import Store

    database, research_id, report_path = _seed_history(tmp_path)
    _write_json_report(report_path)
    _seed_evidence(database, research_id)
    store = Store(database)
    store.replace_report_tags(research_id, ["会议助手", "口碑", "竞品"], source="agent", created_at="2026-08-28T09:00:00Z")
    text = report_path.read_text(encoding="utf-8")
    return export_excel(store, research_id, tmp_path / "runs", text), research_id


def test_导出六_sheet_且八项校验全过(tmp_path: Path) -> None:
    from app.export.excel_check import EXPECTED_SHEETS, check_workbook

    path, research_id = _export(tmp_path)
    assert path == tmp_path / "runs" / research_id / "exports" / f"{research_id}.xlsx"
    assert check_workbook(path) == []
    wb = load_workbook(path)
    assert wb.sheetnames == EXPECTED_SHEETS and wb["90_图表数据"].sheet_state == "hidden"
    sources = wb["04_信息源"]
    assert [sources.cell(row=r, column=1).value for r in range(2, 4)] == ["S01", "S02"]
    assert sources.cell(row=4, column=1).value is None  # 只列 citation_no 非空的行
    assert [sources.cell(row=2, column=c).value for c in range(6, 12)] == ["? / ?", 1, 2, None, 1, 1]
    assert sources["D2"].hyperlink.target == "https://www.xiaohongshu.com/explore/a1"
    assert "图例" in str(wb["01_结论摘要"]["A3"].value) and "? / ?" in str(wb["01_结论摘要"]["A3"].value)
    assert wb["01_结论摘要"]["A4"].value == "1. 结论一 [S01][S02]"
    assert wb["03_明细数据"].max_row == 5 and wb["03_明细数据"]["J3"].value == 3  # 原始:digg_count
    assert [wb["05_标签"].cell(row=r, column=1).value for r in (2, 3, 4)] == ["会议助手", "口碑", "竞品"]
    charts = [n for n in zipfile.ZipFile(path).namelist() if n.startswith("xl/charts/chart")]
    assert len(charts) >= 2


def test_校验器能抓出故意做坏的产物(tmp_path: Path) -> None:
    from app.export.excel_check import check_workbook

    path, _ = _export(tmp_path)
    wb = load_workbook(path)
    wb.remove(wb["05_标签"])
    wb.save(tmp_path / "nosheet.xlsx")
    assert any(e.startswith("1 ") for e in check_workbook(tmp_path / "nosheet.xlsx"))

    wb = load_workbook(path)
    wb["04_信息源"]["D2"].hyperlink = None
    wb["04_信息源"]["G3"] = 5
    wb["04_信息源"]["F2"] = "3 / D"
    wb["01_结论摘要"]["A5"] = "2. 结论没角标"
    wb["90_图表数据"].sheet_state = "visible"
    wb.save(tmp_path / "cells.xlsx")
    codes = sorted({e.split(" ", 1)[0] for e in check_workbook(tmp_path / "cells.xlsx")})
    assert codes == ["1", "3", "5", "7"]


def test_m6e_raw_metrics_带嵌套_raw_也导得出(tmp_path: Path) -> None:
    """§M6-e 关账整跑实证：weibo/wechat_mp/web_search 的 raw_metrics 里带一个
    `_raw` 子字典（契约「解析值+原始串双留」），摊成「原始:_raw」列后 openpyxl
    直接 `Cannot convert {...} to Excel`，**整份导出 500**——判据②与⑤d 两条红
    都是它一个引起的（DLV-1 当时底料只有 xhs，没有 `_raw`，所以没踩到）。

    修两层：`_raw` 不进指标列；单元格再兜一道非标量转 JSON 串。
    """
    import json as _json
    import sqlite3

    from app.export.excel import export_excel
    from app.export.excel_check import check_workbook
    from app.store.dao import Store
    from tests.test_dlv1_delivery import _seed_evidence, _seed_history, _write_json_report

    database, research_id, report_path = _seed_history(tmp_path)
    _write_json_report(report_path)
    _seed_evidence(database, research_id)
    # 把真跑里 wechat_mp 的形状原样塞进去（含嵌套 _raw 与全 None）
    con = sqlite3.connect(database)
    con.execute(
        "UPDATE evidence SET platform='wechat_mp', raw_metrics=? WHERE id=("
        "SELECT id FROM evidence LIMIT 1)",
        (_json.dumps({"read_count": None, "like_count": None,
                      "_raw": {"read_count": None, "like_count": None}}),),
    )
    con.commit()
    con.close()

    path = export_excel(Store(database), research_id, tmp_path / "runs",
                        report_path.read_text(encoding="utf-8"))
    assert path.is_file() and check_workbook(path) == []
    headers = [c.value for c in load_workbook(path)["03_明细数据"][2]]
    assert not any(str(h or "").startswith("原始:_raw") for h in headers)
