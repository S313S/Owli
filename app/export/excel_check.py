"""Excel 附件产物校验器（spec §6 八项，§DLV-1 货 3）。

闭集写死在本文件，**不 import 生成器常量**——尺子与被量的东西分开。
降级规则（本包自拍，worklog §3）：H–L 某格为空 = 该维不可评（库里 NULL、
rating_notes 记 `?`），此时 G 列须为 `? / ?`、M 列该段须为 `?`；五维全空且无 M 视为未评级。
五维齐全时按 spec 严格断言 G = 和、等级阈值、M 各段 = H–L。
§CMT-1 货 5：04_信息源 C 列插入「类型（帖/评论）」，其后各列整体右移一位。
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

EXPECTED_SHEETS = ["01_结论摘要", "02_图表", "03_明细数据", "04_信息源", "05_标签", "90_图表数据"]
MARK = re.compile(r"\[S\d{2}\]")
MARK_ID = re.compile(r"\bS\d{2}\b")
# §RATE-4 货 1：UGC 行第一维叫「代表性」，尺子跟着放宽（口径的唯一定义在
# `app.reliability.scoring.RATING_NOTES_PATTERN`，这里是导出侧的独立复核）。
NOTES = re.compile(
    r"^(?:权威|代表性)([0-2?]):(.{1,14}) · 时效([0-2?]):(.{1,14}) · "
    r"交叉([0-2?]):(.{1,14}) · 完整([0-2?]):(.{1,14}) · "
    r"无关([0-2?]):(.{1,14})( ⚠️.{1,30})?$"
)
SUMMARY = re.compile(r"^(\d{1,2}|\?) / ([ABCD?])$")


def _grade(total: int) -> str:
    return "A" if total >= 8 else "B" if total >= 6 else "C" if total >= 4 else "D"


def _conclusion_rows(ws) -> list[str]:
    """`01` 从第 4 行起（第 3 行是图例，§AUTO-EXP 货 6）、A 列连续非空的行 = 结论行。"""
    rows: list[str] = []
    for r in range(4, ws.max_row + 1):
        value = ws.cell(row=r, column=1).value
        if value is None or str(value).strip() == "":
            break
        rows.append(str(value))
    return rows


def check_workbook(path: str | Path) -> list[str]:
    """逐条跑 spec §6 八项；返回错误列表，空列表 = 全过。"""
    errors: list[str] = []
    try:
        wb = load_workbook(path)  # 第 6 项：能被 openpyxl 打开（zip 完整）
    except Exception as error:  # noqa: BLE001 — 任何打不开都算第 6 项失败
        return [f"6 无法打开: {error}"]
    if wb.sheetnames != EXPECTED_SHEETS:
        errors.append(f"1 sheet 顺序/命名不符: {wb.sheetnames}")
        return errors
    if wb["90_图表数据"].sheet_state != "hidden":
        errors.append("1 90_图表数据 未隐藏")
    if len(wb["02_图表"]._charts) < 1:
        errors.append("2 02_图表 无内嵌图表")
    legend = str(wb["01_结论摘要"]["A3"].value or "")
    if "图例" not in legend or "? / ?" not in legend:
        errors.append(f"3 01 第 3 行缺图例（? / ? 记法说明）: {legend[:30]!r}")
    conclusions = _conclusion_rows(wb["01_结论摘要"])
    for text in conclusions:
        if not re.search(r"\[S\d{2}\]\s*$", text):
            errors.append(f"3 结论行末尾无角标: {text[:30]}…")
    cited: set[str] = set()
    for ws in (wb["01_结论摘要"], wb["02_图表"], wb["03_明细数据"]):
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cited.update(MARK_ID.findall(cell.value))
    listed = {str(r[0].value) for r in wb["04_信息源"].iter_rows(min_row=2, max_col=1) if r[0].value}
    if not cited <= listed:
        errors.append(f"4 悬空角标: {sorted(cited - listed)}")
    errors.extend(_check_sources(wb["04_信息源"]))
    return errors


def _check_sources(ws) -> list[str]:
    """第 5/7/8 项：E 列超链接；H–L 0–2 整数、G=和且等级一致；M 与 H–L 对应。

    另加一项（§CMT-1 货 5）：C 列类型只能是「帖」或「评论」。
    """
    errors: list[str] = []
    for r in range(2, ws.max_row + 1):
        mark = ws.cell(row=r, column=1).value
        if not mark:
            continue
        url = ws.cell(row=r, column=5)
        if not url.value or not url.hyperlink:
            errors.append(f"5 行{r} E 列非超链接或为空")
        kind = str(ws.cell(row=r, column=3).value or "")
        if kind not in {"帖", "评论"}:
            errors.append(f"9 行{r} C 列类型不在闭集: {kind!r}")
        scores = [ws.cell(row=r, column=c).value for c in range(8, 13)]
        summary = str(ws.cell(row=r, column=7).value or "")
        notes = str(ws.cell(row=r, column=13).value or "")
        bad = [s for s in scores if s is not None and (not isinstance(s, int) or isinstance(s, bool) or not 0 <= s <= 2)]
        if bad:
            errors.append(f"7 行{r} H–L 含非 0–2 整数: {scores}")
            continue
        if any(s is None for s in scores):  # 某维不可评（`?` 记法）：F 列须为 `? / ?`，不重算
            if summary != "? / ?":
                errors.append(f"7 行{r} 有维度未评但 G 列不是 '? / ?': {summary}")
            if all(s is None for s in scores) and not notes:
                continue  # 完全未评级：M 允许为空
        else:
            total = sum(scores)
            matched = SUMMARY.match(summary)
            if not matched or matched.group(1) != str(total) or matched.group(2) != _grade(total):
                errors.append(f"7 行{r} G 列与 H–L 不一致: {summary!r} vs {total}/{_grade(total)}")
        parsed = NOTES.match(notes)
        if not parsed:
            errors.append(f"8 行{r} M 列不合 rating_notes 正则: {notes[:40]!r}")
        elif [parsed.group(i) for i in (1, 3, 5, 7, 9)] != ["?" if s is None else str(s) for s in scores]:
            errors.append(f"8 行{r} M 列各段分数与 H–L 不等")
    return errors


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print("用法: python -m app.export.excel_check <file.xlsx>")
        return 2
    errors = check_workbook(argv[1])
    print("\n".join(errors) if errors else f"✅ 八项全过: {argv[1]}")
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
