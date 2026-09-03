"""Excel 附件确定性导出（§DLV-1 货 3）。

严格按 `docs/design/report-attachment-spec.md` §2/§5/§6：6 sheet 顺序命名不可变、
`90` 隐藏、`04_信息源` 12 列定死、G–K 数值型 + 条件格式、D 列超链接。
只读库里的评分，不重算；agent 选图不在本包（只画两张标准原生图）。
"""

from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.reliability.scoring import SCORE_FIELDS, grade_for_total

SHEETS = ("01_结论摘要", "02_图表", "03_明细数据", "04_信息源", "05_标签", "90_图表数据")
SOURCE_HEADER = (
    # §CMT-1 货 5：C 列「类型」区分帖/评论——读者反应和帖子作者的说法
    # 权重不一样，人工复核第一眼要能分开。其后各列整体右移一位。
    "角标ID", "平台", "类型", "标题/摘要", "URL", "抓取时间", "总分与等级",
    "权威性(0-2)", "时效性(0-2)", "交叉验证(0-2)", "完整度(0-2)", "利益无关性(0-2)",
    "评分理由与备注",
)
#: kind 列到中文表头的映射；旧库没有 kind 列的行按帖处理。
KIND_LABELS = {"post": "帖", "comment": "评论"}
C_DEEP, C_MID, C_DGRAY = "1F3864", "4472C4", "595959"
F_TITLE = Font(name="微软雅黑", size=12, bold=True, color=C_DEEP)
F_SUB = Font(name="微软雅黑", size=9, color=C_DGRAY)
F_HEAD = Font(name="微软雅黑", size=10, bold=True)
F_BODY = Font(name="微软雅黑", size=10)
F_TLDR = Font(name="微软雅黑", size=11, bold=True)
F_LINK = Font(name="Calibri", size=10, color="0563C1", underline="single")
FILL_ZERO, FILL_TWO = PatternFill("solid", fgColor="FCE4D6"), PatternFill("solid", fgColor="D9E1F2")
_MARK = re.compile(r"\[S\d{2}\]")


def _mark(number: int) -> str:
    return f"S{number:02d}"


def _cellable(value: Any) -> Any:
    """§M6-e：openpyxl 只收标量，塞进去一个 dict/list 就 `Cannot convert … to Excel`
    并把**整份导出**掀掉（500）。raw_metrics 的契约本来就允许嵌套（`_raw` 子字典），
    所以这里兜一道：非标量一律转成紧凑 JSON 串，宁可单元格难看也不能整表导不出。"""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (Mapping, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


def _write_rows(ws: Worksheet, start: int, rows: Sequence[Sequence[Any]], *, header: bool = True) -> int:
    for r, row in enumerate(rows, start=start):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=_cellable(value))
            cell.font = F_HEAD if (header and r == start) else F_BODY
    return start + len(rows)


def _fetched(value: Any) -> str:
    """spec §5 E 列：`YYYY-MM-DD HH:MM`（UTC+8）；解析不了原样留。"""
    from datetime import datetime, timedelta, timezone

    text = str(value or "")
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return text
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone(timedelta(hours=8))).strftime("%Y-%m-%d %H:%M")


def _sheet_sources(ws: Worksheet, cited: Sequence[Mapping[str, Any]]) -> None:
    """`04_信息源`：spec §5 12 列定死；未评级维度留空、F 列记 `?`（不重算）。"""
    rows: list[list[Any]] = [list(SOURCE_HEADER)]
    for item in cited:
        scores = [item.get(field) for field in SCORE_FIELDS]
        complete = all(isinstance(s, int) for s in scores)
        total = sum(scores) if complete else None
        summary = f"{total} / {grade_for_total(total)}" if complete else "? / ?"
        rows.append([
            _mark(int(item["citation_no"])), item.get("platform"),
            KIND_LABELS.get(str(item.get("kind") or "post"), "帖"),
            item.get("title") or item.get("content_excerpt") or item.get("permalink"),
            item.get("permalink"), _fetched(item.get("fetched_at")), summary,
            *[s if isinstance(s, int) else None for s in scores],
            item.get("rating_notes") or "",
        ])
    last = _write_rows(ws, 1, rows)
    for r in range(2, last):
        cell = ws.cell(row=r, column=5)
        cell.hyperlink = str(cell.value)
        cell.font = F_LINK
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
        for c in range(8, 13):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
    if last > 2:
        span = f"H2:L{last - 1}"
        ws.conditional_formatting.add(span, CellIsRule(operator="equal", formula=["0"], fill=FILL_ZERO))
        ws.conditional_formatting.add(span, CellIsRule(operator="equal", formula=["2"], fill=FILL_TWO))
    for col, width in zip("ABCDEFGHIJKLM", (8, 12, 6, 48, 40, 17, 12, 8, 8, 8, 8, 8, 60)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"


def _sheet_summary(ws: Worksheet, question: str, conclusions: Sequence[str], title: str) -> None:
    """`01_结论摘要`：A1 调研问题、A2 计数行、A3 图例行（§AUTO-EXP 货 6）、A4 起每条结论
    （末尾须带角标，spec §5.2）。"""
    ws["A1"] = f"调研问题：{question}"
    ws["A1"].font = F_TITLE
    ws["A2"] = f"报告：{title} · 结论 {len(conclusions)} 条（从成稿「结论」段确定性摘取）"
    ws["A2"].font = F_SUB
    ws["A3"] = "图例：04_信息源 G–K 某格留空 = 该维不可评（不重算）；此时 F 列总分记「? / ?」"
    ws["A3"].font = F_SUB
    for i, text in enumerate(conclusions):
        body = text.strip()
        marks = "".join(_MARK.findall(body))
        stripped = _MARK.sub("", body).strip()
        cell = ws.cell(row=4 + i, column=1, value=f"{i + 1}. {stripped} {marks}".rstrip())
        cell.font = F_TLDR
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110


def _sheet_details(ws: Worksheet, evidence: Sequence[Mapping[str, Any]], marks: str) -> None:
    """`03_明细数据`：全部证据摊平；原始指标列按平台原样 + 平台内归一化列（R4）。"""
    # `_raw` 是 raw_metrics 契约里的「原始串镜像」，逐键重复已解析值，摊成一列只会得到
    # 一列字典（M6 的 weibo/wechat_mp/web_search 都带它）——不作指标列。
    metric_keys = sorted({
        str(k) for item in evidence
        for k in (item.get("raw_metrics") or {}) if isinstance(item.get("raw_metrics"), Mapping)
    } - {"_raw"})
    ws["A1"] = f"表1 证据明细（{len(evidence)} 条，原始指标不可跨平台相加） {marks}".rstrip()
    ws["A1"].font = F_HEAD
    header = ["角标ID", "平台", "类型", "采集方式", "标题", "URL", "作者", "发布时间", "抓取时间",
              *[f"原始:{k}" for k in metric_keys], "归一化分", "归一化方法", "等级"]
    rows: list[list[Any]] = [header]
    for item in evidence:
        metrics = item.get("raw_metrics") if isinstance(item.get("raw_metrics"), Mapping) else {}
        rows.append([
            _mark(item["citation_no"]) if item.get("citation_no") is not None else None,
            item.get("platform"), item.get("source_type"), item.get("fetch_method"),
            item.get("title") or item.get("content_excerpt"), item.get("permalink"),
            item.get("author_name"), item.get("published_at"), _fetched(item.get("fetched_at")),
            *[metrics.get(k) for k in metric_keys],
            item.get("normalized_score"), item.get("norm_method"), item.get("grade"),
        ])
    _write_rows(ws, 2, rows)
    ws.freeze_panes = "A3"
    for col, width in zip("ABCDEFGHI", (8, 10, 8, 14, 48, 40, 14, 20, 17)):
        ws.column_dimensions[col].width = width


def _sheet_tags(ws: Worksheet, tags: Sequence[str]) -> None:
    """`05_标签`：agent 现值；空则只表头（spec §2 行 5）。"""
    rows: list[list[Any]] = [["标签名", "类型", "判定依据", "关联结论编号"]]
    rows.extend([tag, "agent", "report_tags 现值", ""] for tag in tags)
    _write_rows(ws, 1, rows)
    for col, width in zip("ABCD", (24, 10, 30, 14)):
        ws.column_dimensions[col].width = width


def _count_table(ws: Worksheet, start: int, label: str, header: tuple[str, str],
                 counts: Counter[str]) -> tuple[int, int]:
    """向 `90` 写一张两列计数表，返回 (表头行, 末行)。"""
    ws.cell(row=start, column=1, value=label).font = F_HEAD
    rows: list[list[Any]] = [list(header)]
    rows.extend([key, count] for key, count in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0])))
    end = _write_rows(ws, start + 1, rows) - 1
    return start + 1, end


def _bar(ws_stage: Worksheet, head: int, end: int, color: str) -> BarChart:
    chart = BarChart()
    chart.type = "col"
    data = Reference(ws_stage, min_col=2, min_row=head, max_row=end)
    cats = Reference(ws_stage, min_col=1, min_row=head + 1, max_row=end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    series = chart.series[0]
    series.graphicalProperties.solidFill = color
    series.dLbls = DataLabelList(showVal=True, showSerName=False, showCatName=False, showLegendKey=False)
    chart.legend = None
    chart.title = None
    chart.height, chart.width = 11, 24
    return chart


def _sheet_charts(ws: Worksheet, ws_stage: Worksheet, evidence: Sequence[Mapping[str, Any]],
                  cited_marks: str) -> None:
    """`02_图表`：两张标准原生图（平台分布 / 等级分布），数据在 `90`（spec §2 行 2、§4.5）。"""
    ws.column_dimensions["A"].width = 110
    platforms = Counter(str(item.get("platform")) for item in evidence)
    grades = Counter(str(item.get("grade") or "未评级") for item in evidence)
    p_head, p_end = _count_table(ws_stage, 1, "C1 证据平台分布", ("平台", "证据数"), platforms)
    g_head, g_end = _count_table(ws_stage, p_end + 3, "C2 证据可靠度等级分布", ("等级", "证据数"), grades)
    top_platform, top_n = (platforms.most_common(1) or [("—", 0)])[0]
    blocks = [
        (f"证据主要来自「{top_platform}」（{top_n}/{len(evidence)} 条）", f"条 · 全部证据 · n={len(evidence)}",
         f"来源: {cited_marks or '—'} · 数据: 90_图表数据!A{p_head}:B{p_end}", _bar(ws_stage, p_head, p_end, C_MID)),
        (f"可靠度等级分布：{'、'.join(f'{k} {v}' for k, v in grades.most_common(4))}",
         f"条 · 按 evidence.grade · n={len(evidence)}",
         f"来源: {cited_marks or '—'} · 数据: 90_图表数据!A{g_head}:B{g_end}", _bar(ws_stage, g_head, g_end, C_DEEP)),
    ]
    for index, (title, sub, source_line, chart) in enumerate(blocks):
        row = 1 + index * 28
        ws.cell(row=row, column=1, value=title).font = F_TITLE
        ws.cell(row=row + 1, column=1, value=sub).font = F_SUB
        ws.cell(row=row + 25, column=1, value=source_line).font = F_SUB
        ws.add_chart(chart, f"A{row + 2}")


def build_workbook(report: Mapping[str, Any], view: Mapping[str, Any],
                   evidence: Sequence[Mapping[str, Any]], tags: Sequence[str]) -> Workbook:
    """报告行 + 结构化视图 + 证据 + 标签 → 6 sheet 工作簿（纯内存，确定性）。"""
    wb = Workbook()
    wb.remove(wb.active)
    sheets = {name: wb.create_sheet(name) for name in SHEETS}
    sheets["90_图表数据"].sheet_state = "hidden"
    cited = sorted(
        (item for item in evidence if item.get("citation_no") is not None),
        key=lambda item: int(item["citation_no"]),
    )
    cited_marks = ", ".join(_mark(int(item["citation_no"])) for item in cited)
    bracketed = "".join(f"[{_mark(int(item['citation_no']))}]" for item in cited)
    _sheet_summary(sheets["01_结论摘要"], str(report.get("research_question") or ""),
                   list(view.get("conclusions") or []), str(view.get("title") or report.get("title") or ""))
    _sheet_charts(sheets["02_图表"], sheets["90_图表数据"], evidence, cited_marks)
    _sheet_details(sheets["03_明细数据"], evidence, bracketed)
    _sheet_sources(sheets["04_信息源"], cited)
    _sheet_tags(sheets["05_标签"], tags)
    return wb


def export_excel(store: Any, research_id: str, runs_root: Path, report_text: str) -> Path:
    """落盘 `runs/<id>/exports/<id>.xlsx` 并返回路径；登记交给 registry（货 5）。"""
    from app.report.render import parse_report

    report = store.get_report(research_id)
    if report is None:
        raise KeyError(f"报告不存在：{research_id}")
    view = parse_report(report_text)
    evidence = store.list_evidence(research_id)
    tags = store.read_validation_path("report_tags", research_id)
    target = Path(runs_root) / research_id / "exports" / f"{research_id}.xlsx"
    target.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(report, view, evidence, tags).save(target)
    return target
